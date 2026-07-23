"""Build a cached fundamentals snapshot.

Used by the GitHub Action (and importable from app.py for the download button).
"""

from __future__ import annotations

import argparse
import io
from pathlib import Path

import pandas as pd

from metrics import build_report


REPO_ROOT = Path(__file__).resolve().parent
DEFAULT_TICKERS_CSV = REPO_ROOT / "tickers.csv"
DEFAULT_PARQUET = REPO_ROOT / "snapshot.parquet"
DEFAULT_XLSX = REPO_ROOT / "report.xlsx"


# Per-column display config used by the Excel writer.
# (header label, openpyxl number format)
EXCEL_FORMATS: dict[str, tuple[str, str]] = {
    "ticker": ("Ticker", "@"),
    "benchmark": ("Benchmark", "@"),
    "note": ("Note", "@"),
    "sector": ("Sector", "@"),
    "industry": ("Industry", "@"),
    "last_price": ("Last Price", "#,##0.00"),
    "ytd_pct": ("YTD %", "0.00%"),
    "pct_from_52w_high": ("% from 52w High", "0.00%"),
    "fifty_two_week_high": ("52w High", "#,##0.00"),
    "fifty_two_week_low": ("52w Low", "#,##0.00"),
    "market_cap": ("Market Cap", "#,##0"),
    "pe_trailing": ("P/E (TTM)", "0.00"),
    "pe_forward": ("P/E (Fwd)", "0.00"),
    "ps": ("P/S", "0.00"),
    "pb": ("P/B", "0.00"),
    "peg": ("PEG", "0.00"),
    "ev_ebitda": ("EV/EBITDA", "0.00"),
    "ev_sales": ("EV/Sales", "0.00"),
    "fcf_yield": ("FCF Yield", "0.00%"),
    "rev_growth_ttm": ("Rev Growth (TTM)", "0.00%"),
    "fwd_eps_growth": ("Fwd EPS Growth", "0.00%"),
    "fwd_rev_growth": ("Fwd Rev Growth", "0.00%"),
    "roa": ("ROA", "0.00%"),
    "roe": ("ROE", "0.00%"),
    "gross_margin": ("Gross Margin", "0.00%"),
    "operating_margin": ("Operating Margin", "0.00%"),
    "free_cash_flow": ("Free Cash Flow", "#,##0"),
    "current_ratio": ("Current Ratio", "0.00"),
    "quick_ratio": ("Quick Ratio", "0.00"),
    "debt_equity": ("Debt/Equity", "0.00"),
    "debt_assets": ("Debt/Assets", "0.00"),
    "net_debt_ebitda": ("Net Debt/EBITDA", "0.00"),
    "interest_coverage": ("Interest Coverage", "0.00"),
    "dividend_yield": ("Dividend Yield", "0.00%"),
    "beta": ("Beta", "0.00"),
    "vol_2m_annualized": ("Vol 2m (ann.)", "0.00%"),
    "avg_dollar_vol_30d": ("Avg $ Vol (30d)", "#,##0"),
    "next_earnings": ("Next Earnings", "yyyy-mm-dd"),
    "missing_fields": ("Missing", "@"),
    "as_of": ("As Of", "@"),
}


def write_excel(df: pd.DataFrame, path: str | Path) -> None:
    """Write a formatted xlsx with a frozen header row and column widths."""
    path = Path(path)
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name="Fundamentals")
        _format_sheet(xw, df, sheet_name="Fundamentals")


def build_excel_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name="Fundamentals")
        _format_sheet(xw, df, sheet_name="Fundamentals")
    return buf.getvalue()


def _format_sheet(xw: pd.ExcelWriter, df: pd.DataFrame, *, sheet_name: str) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = xw.sheets[sheet_name]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_align = Alignment(horizontal="center", vertical="center")

    # Header row: rename to friendly labels + style.
    for col_idx, col_name in enumerate(df.columns, start=1):
        label, _ = EXCEL_FORMATS.get(col_name, (col_name, "@"))
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Number formats + column widths.
    for col_idx, col_name in enumerate(df.columns, start=1):
        _, fmt = EXCEL_FORMATS.get(col_name, (col_name, "General"))
        for row_idx in range(2, len(df) + 2):
            ws.cell(row=row_idx, column=col_idx).number_format = fmt
        label = EXCEL_FORMATS.get(col_name, (col_name,))[0]
        sample_lens = [len(str(label))]
        sample_lens += [len(str(v)) for v in df[col_name].head(20).tolist() if v is not None]
        width = min(max(sample_lens + [10]) + 2, 28)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.freeze_panes = "A2"


def load_tickers(path: str | Path = DEFAULT_TICKERS_CSV) -> pd.DataFrame:
    return pd.read_csv(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Build fundamentals snapshot.")
    parser.add_argument("--tickers", default=str(DEFAULT_TICKERS_CSV))
    parser.add_argument("--parquet", default=str(DEFAULT_PARQUET))
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    parser.add_argument("--delay", type=float, default=0.25)
    args = parser.parse_args()

    tickers_df = load_tickers(args.tickers)
    print(f"Building report for {len(tickers_df)} tickers...")
    report = build_report(tickers_df, delay=args.delay)
    report.to_parquet(args.parquet, index=False)
    write_excel(report, args.xlsx)
    print(f"Wrote {args.parquet} and {args.xlsx} ({len(report)} rows)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
